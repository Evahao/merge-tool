import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

class MergeValidator:
    def __init__(self, root):
        self.root = root
        self.root.title("高级表合并校验工具（单价冲突版）")
        self.root.geometry("800x650")
        
        self.file_a_path = ""
        self.file_b_path = ""
        self.df_a = None
        self.df_b = None
        self.sheet_a = None
        self.sheet_b = None
        
        self.output_dir = tk.StringVar(value=os.getcwd())
        self.mapping_columns = []
        
        self.key_column = "客户料号"
        self.price_column = "单价"
        
        self.config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mapping_config.json")
        self.create_widgets()
        self.load_config()

    def create_widgets(self):
        tk.Label(self.root, text="表合并校验工具（仅校验客户料号单价冲突）", font=("Arial", 14, "bold")).pack(pady=5)
        
        frame_files = tk.Frame(self.root)
        frame_files.pack(fill="x", padx=20, pady=5)
        
        frame_a = tk.LabelFrame(frame_files, text="表A", padx=5, pady=5)
        frame_a.pack(side="left", fill="both", expand=True, padx=5)
        self.label_a = tk.Label(frame_a, text="未选择", fg="gray")
        self.label_a.pack(anchor="w")
        tk.Button(frame_a, text="浏览", command=self.load_a).pack(anchor="w", pady=2)
        
        frame_b = tk.LabelFrame(frame_files, text="表B", padx=5, pady=5)
        frame_b.pack(side="right", fill="both", expand=True, padx=5)
        self.label_b = tk.Label(frame_b, text="未选择", fg="gray")
        self.label_b.pack(anchor="w")
        tk.Button(frame_b, text="浏览", command=self.load_b).pack(anchor="w", pady=2)
        
        frame_output = tk.LabelFrame(self.root, text="输出目录", padx=10, pady=5)
        frame_output.pack(fill="x", padx=20, pady=5)
        self.output_label = tk.Label(frame_output, textvariable=self.output_dir, fg="blue", anchor="w")
        self.output_label.pack(side="left", fill="x", expand=True)
        tk.Button(frame_output, text="更改目录", command=self.select_output_dir).pack(side="right")
        
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="配置字段映射", command=self.open_mapping_dialog, bg="#3498db", fg="white", width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="保存配置", command=self.save_config, bg="#2ecc71", fg="white", width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="加载配置", command=self.load_config, bg="#f39c12", fg="white", width=15).pack(side="left", padx=5)
        
        self.process_btn = tk.Button(self.root, text="合并并校验单价", command=self.process, bg="#e67e22", fg="white", font=("Arial", 12), width=20)
        self.process_btn.pack(pady=15)
        
        self.status = tk.Label(self.root, text="就绪", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

    def select_output_dir(self):
        dir_selected = filedialog.askdirectory(title="选择输出目录")
        if dir_selected:
            self.output_dir.set(dir_selected)

    def load_a(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.file_a_path = path
            try:
                self.sheet_a = self.select_sheet(path, "选择表A的工作表")
                self.df_a = pd.read_excel(path, sheet_name=self.sheet_a)
                self.label_a.config(text=f"{os.path.basename(path)} (工作表: {self.sheet_a})", fg="black")
                self.status.config(text=f"表A加载成功，共{len(self.df_a)}行")
            except Exception as e:
                messagebox.showerror("错误", f"加载表A失败：{e}")
                self.df_a = None
                self.label_a.config(text="加载失败", fg="red")

    def load_b(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.file_b_path = path
            try:
                self.sheet_b = self.select_sheet(path, "选择表B的工作表")
                self.df_b = pd.read_excel(path, sheet_name=self.sheet_b)
                self.label_b.config(text=f"{os.path.basename(path)} (工作表: {self.sheet_b})", fg="black")
                self.status.config(text=f"表B加载成功，共{len(self.df_b)}行")
            except Exception as e:
                messagebox.showerror("错误", f"加载表B失败：{e}")
                self.df_b = None
                self.label_b.config(text="加载失败", fg="red")

    def select_sheet(self, file_path, title):
        xl = pd.ExcelFile(file_path)
        sheets = xl.sheet_names
        if len(sheets) == 1:
            return sheets[0]
        else:
            dialog = tk.Toplevel(self.root)
            dialog.title(title)
            dialog.geometry("300x150")
            dialog.transient(self.root)
            dialog.grab_set()
            tk.Label(dialog, text="该文件包含多个工作表，请选择：").pack(pady=10)
            var = tk.StringVar()
            combo = ttk.Combobox(dialog, textvariable=var, values=sheets, state="readonly", width=30)
            combo.pack(pady=10)
            combo.current(0)
            result = [None]
            def on_ok():
                result[0] = var.get()
                dialog.destroy()
            tk.Button(dialog, text="确定", command=on_ok, bg="#4CAF50", fg="white").pack(pady=10)
            self.root.wait_window(dialog)
            if result[0] is None:
                raise ValueError("未选择工作表")
            return result[0]

    def open_mapping_dialog(self):
        if self.df_a is None or self.df_b is None:
            messagebox.showerror("错误", "请先加载表A和表B")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("字段映射配置")
        dialog.geometry("850x550")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="默认固定列：客户料号、HQ料号、Category、Usag、单价、Extend（客户料号不可删除）", font=("Arial", 10)).pack(pady=5)
        
        header_frame = tk.Frame(dialog)
        header_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(header_frame, text="导出列名", width=18, font=("Arial", 9, "bold")).pack(side="left", padx=2)
        tk.Label(header_frame, text="表A实际列", width=25, font=("Arial", 9, "bold")).pack(side="left", padx=2)
        tk.Label(header_frame, text="表B实际列", width=25, font=("Arial", 9, "bold")).pack(side="left", padx=2)
        tk.Label(header_frame, text="操作", width=10, font=("Arial", 9, "bold")).pack(side="left", padx=2)
        
        self.rows = []
        canvas = tk.Canvas(dialog, height=320)
        scrollbar = tk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=5)
        scrollbar.pack(side="right", fill="y")
        
        cols_a = [""] + list(self.df_a.columns)
        cols_b = [""] + list(self.df_b.columns)
        
        default_cols = [
            {"export_name": "客户料号", "col_a": "", "col_b": ""},
            {"export_name": "HQ料号", "col_a": "", "col_b": ""},
            {"export_name": "Category", "col_a": "", "col_b": ""},
            {"export_name": "Usag", "col_a": "", "col_b": ""},
            {"export_name": "单价", "col_a": "", "col_b": ""},
            {"export_name": "Extend", "col_a": "", "col_b": ""}
        ]
        
        if not self.mapping_columns:
            self.mapping_columns = default_cols
        
        for col in self.mapping_columns:
            self.add_mapping_row(scrollable_frame, col["export_name"], col["col_a"], col["col_b"], cols_a, cols_b)
        
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="添加一行", command=lambda: self.add_mapping_row(scrollable_frame, "", "", "", cols_a, cols_b), bg="#3498db", fg="white").pack(side="left", padx=5)
        tk.Button(btn_frame, text="确定保存", command=lambda: self.save_mapping_from_dialog(dialog), bg="#2ecc71", fg="white").pack(side="left", padx=5)
        tk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side="left", padx=5)

    def add_mapping_row(self, parent, export_name="", col_a="", col_b="", cols_a=None, cols_b=None):
        row_frame = tk.Frame(parent)
        row_frame.pack(fill="x", pady=2)
        
        entry_export = tk.Entry(row_frame, width=18)
        entry_export.insert(0, export_name)
        entry_export.pack(side="left", padx=2)
        
        combo_a = ttk.Combobox(row_frame, values=cols_a, width=25)
        combo_a.set(col_a)
        combo_a.pack(side="left", padx=2)
        
        combo_b = ttk.Combobox(row_frame, values=cols_b, width=25)
        combo_b.set(col_b)
        combo_b.pack(side="left", padx=2)
        
        def remove():
            if export_name.strip() == "客户料号":
                messagebox.showwarning("提示", "客户料号为必填字段，不可删除！")
                return
            row_frame.destroy()
        
        if export_name.strip() == "客户料号":
            del_btn = tk.Button(row_frame, text="必填", state="disabled", width=8)
        else:
            del_btn = tk.Button(row_frame, text="删除", command=remove, fg="red", width=8)
        del_btn.pack(side="left", padx=2)
        self.rows.append((row_frame, entry_export, combo_a, combo_b))

    def save_mapping_from_dialog(self, dialog):
        new_mapping = []
        has_key = False
        for row_frame, entry_export, combo_a, combo_b in self.rows:
            export_name = entry_export.get().strip()
            if not export_name:
                continue
            if export_name == "客户料号":
                has_key = True
            col_a = combo_a.get().strip()
            col_b = combo_b.get().strip()
            new_mapping.append({
                "export_name": export_name,
                "col_a": col_a if col_a else None,
                "col_b": col_b if col_b else None
            })
        if not has_key:
            messagebox.showerror("错误", "客户料号为必填项，必须保留！")
            return
        if not new_mapping:
            messagebox.showerror("错误", "至少需要定义一个导出列")
            return
        self.mapping_columns = new_mapping
        dialog.destroy()
        self.status.config(text=f"映射已更新，共{len(self.mapping_columns)}列")

    def save_config(self):
        config = {
            "mapping_columns": self.mapping_columns,
            "output_dir": self.output_dir.get()
        }
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            self.status.config(text="配置已保存")
            messagebox.showinfo("成功", "映射配置已保存")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{e}")

    def load_config(self):
        if not os.path.exists(self.config_path):
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                config = json.load(f)
            if "mapping_columns" in config:
                self.mapping_columns = config["mapping_columns"]
                self.status.config(text="已加载历史映射配置")
            if "output_dir" in config and os.path.exists(config["output_dir"]):
                self.output_dir.set(config["output_dir"])
        except Exception as e:
            messagebox.showerror("错误", f"加载配置失败：{e}")

    def process(self):
        if self.df_a is None or self.df_b is None:
            messagebox.showerror("错误", "请先加载表A和表B")
            return
        if not self.mapping_columns:
            messagebox.showerror("错误", "请先配置字段映射")
            return
        
        key_ok = False
        price_ok = False
        for c in self.mapping_columns:
            if c["export_name"] == self.key_column:
                if c["col_a"] or c["col_b"]:
                    key_ok = True
            if c["export_name"] == self.price_column:
                if c["col_a"] or c["col_b"]:
                    price_ok = True
        if not key_ok:
            messagebox.showerror("错误", "客户料号必须映射有效列")
            return
        if not price_ok:
            messagebox.showerror("错误", "单价必须映射有效列")
            return
        
        self.status.config(text="正在处理数据...")
        self.root.update()
        
        try:
            df_a_std = self.standardize(self.df_a, "a")
            df_b_std = self.standardize(self.df_b, "b")
            df_merged = pd.concat([df_a_std, df_b_std], ignore_index=True)
            
            df_merged[self.price_column] = pd.to_numeric(df_merged[self.price_column], errors="coerce")
            
            a_name = os.path.splitext(os.path.basename(self.file_a_path))[0]
            b_name = os.path.splitext(os.path.basename(self.file_b_path))[0]
            a_sheet = self.sheet_a
            b_sheet = self.sheet_b
            
            df_a_final = df_merged[df_merged["数据来源"] == f"{a_name}({a_sheet})"]
            df_b_final = df_merged[df_merged["数据来源"] == f"{b_name}({b_sheet})"]
            
            a_price_map = df_a_final.dropna(subset=[self.price_column]).groupby(self.key_column)[self.price_column].first().to_dict()
            b_price_map = df_b_final.dropna(subset=[self.price_column]).groupby(self.key_column)[self.price_column].first().to_dict()
            
            conflict_pns = set()
            for pn in a_price_map:
                if pn in b_price_map:
                    p_a = a_price_map[pn]
                    p_b = b_price_map[pn]
                    if abs(p_a - p_b) > 0.0001:
                        conflict_pns.add(pn)
            
            df_merged["是否单价冲突"] = df_merged[self.key_column].isin(conflict_pns)
            
            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            version = "V1.0"
            out_name = f"合并校验结果_{now}_{version}.xlsx"
            out_dir = self.output_dir.get()
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(out_dir, out_name)
            
            df_merged.to_excel(out_path, index=False, engine="openpyxl")
            
            wb = load_workbook(out_path)
            ws = wb.active
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            headers = [cell.value for cell in ws[1]]
            conflict_col = headers.index("是否单价冲突") + 1
            
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=conflict_col).value is True:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
            
            wb.save(out_path)
            
            self.status.config(text=f"处理完成：{out_name}")
            messagebox.showinfo("成功", f"导出完成！\n文件路径：{out_path}\n\n黄色行 = 同一客户料号在A、B表单价不同")
        
        except Exception as e:
            self.status.config(text="处理失败")
            messagebox.showerror("错误", f"处理异常：{str(e)}")

    def standardize(self, df, table):
        out = pd.DataFrame(index=df.index)
        for cfg in self.mapping_columns:
            col = cfg["export_name"]
            src = cfg["col_a"] if table == "a" else cfg["col_b"]
            if src and src in df.columns:
                out[col] = df[src]
            else:
                out[col] = None
        
        # ========== 优化点：来源 = 文件名 + sheet名 ==========
        if table == "a":
            fname = os.path.splitext(os.path.basename(self.file_a_path))[0]
            sheet = self.sheet_a
            out["数据来源"] = f"{fname}({sheet})"
        else:
            fname = os.path.splitext(os.path.basename(self.file_b_path))[0]
            sheet = self.sheet_b
            out["数据来源"] = f"{fname}({sheet})"
        # ==================================================
        
        return out

if __name__ == "__main__":
    root = tk.Tk()
    app = MergeValidator(root)
    root.mainloop()