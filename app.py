import streamlit as st
import pandas as pd
import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------- 页面配置 ----------
st.set_page_config(page_title="高级表合并校验工具", layout="wide")

# ---------- 自定义CSS（现代化卡片式UI）----------
st.markdown("""
<style>
    /* 全局样式 */
    .main > div {
        padding: 1.5rem 2rem;
        max-width: 1600px;
        margin: 0 auto;
    }
    body {
        background-color: #f0f2f6;
    }
    /* 卡片容器 */
    .card {
        background: white;
        border-radius: 24px;
        padding: 1.8rem 2rem;
        margin-bottom: 1.8rem;
        box-shadow: 0 8px 20px rgba(0,0,0,0.02);
        border: 1px solid #f1f5f9;
        transition: box-shadow 0.2s;
    }
    .card:hover {
        box-shadow: 0 12px 28px rgba(0,0,0,0.04);
    }
    /* 标题行 */
    .section-title {
        font-size: 1.3rem;
        font-weight: 600;
        color: #0f172a;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
        gap: 8px;
    }
    .section-title span {
        background: #eef2ff;
        padding: 4px 12px;
        border-radius: 40px;
        font-size: 0.8rem;
        color: #3b82f6;
        margin-left: 10px;
    }
    /* 上传区域双栏 */
    .upload-container {
        display: flex;
        gap: 2rem;
    }
    .upload-col {
        flex: 1;
        background: #fafcff;
        border-radius: 20px;
        padding: 1.2rem 1.5rem;
        border: 1px dashed #cbd5e1;
        transition: border-color 0.2s;
    }
    .upload-col:hover {
        border-color: #3b82f6;
    }
    /* 锁定列徽章 */
    .locked-badge {
        background: #e6f0ff;
        color: #1e4bd2;
        padding: 2px 10px;
        border-radius: 30px;
        font-size: 0.75rem;
        font-weight: 500;
        display: inline-flex;
        align-items: center;
        gap: 4px;
    }
    /* 映射表格 */
    .mapping-table-header {
        display: grid;
        grid-template-columns: 2.5fr 2fr 2fr 0.8fr;
        background: #f8fafc;
        padding: 0.8rem 1.2rem;
        border-radius: 16px;
        font-weight: 600;
        color: #334155;
        margin-bottom: 8px;
        border: 1px solid #e2e8f0;
    }
    .mapping-row {
        display: grid;
        grid-template-columns: 2.5fr 2fr 2fr 0.8fr;
        padding: 0.5rem 1.2rem;
        align-items: center;
        border-bottom: 1px solid #f1f5f9;
    }
    .mapping-row:hover {
        background: #fafcff;
    }
    /* 按钮渐变 */
    .primary-gradient-btn button {
        background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%) !important;
        color: white !important;
        border: none !important;
        font-weight: 600 !important;
        border-radius: 40px !important;
        padding: 0.6rem 2rem !important;
        box-shadow: 0 6px 14px rgba(59,130,246,0.25) !important;
        transition: all 0.2s !important;
    }
    .primary-gradient-btn button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 10px 20px rgba(59,130,246,0.35) !important;
    }
    .secondary-btn button {
        background: white !important;
        color: #1e293b !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 40px !important;
        padding: 0.5rem 1.5rem !important;
        transition: all 0.2s !important;
    }
    .secondary-btn button:hover {
        background: #f8fafc !important;
        border-color: #3b82f6 !important;
    }
    /* 冲突高亮行模拟 (用于预览表格) */
    .conflict-row {
        background-color: #fef9c3 !important;
        border-left: 4px solid #eab308 !important;
    }
    /* 统计徽章 */
    .stats-badge {
        background: white;
        border-radius: 40px;
        padding: 6px 16px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.02);
        border: 1px solid #e2e8f0;
        display: inline-block;
        margin-right: 12px;
    }
    /* 表格容器（横向滚动） */
    .table-wrapper {
        overflow-x: auto;
        border-radius: 16px;
        border: 1px solid #eef2f6;
    }
    /* 调整selectbox样式 */
    div[data-baseweb="select"] {
        border-radius: 12px !important;
    }
    /* 警告提示 */
    .alert-info {
        background: #f0f9ff;
        border-left: 4px solid #0ea5e9;
        padding: 0.8rem 1.2rem;
        border-radius: 12px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ---------- 初始化 session_state ----------
FIXED_COLUMNS = ["客户料号", "单价"]   # 锁定列不可删除
DEFAULT_COLUMNS = ["客户料号", "HQ料号", "Category", "Usage", "单价", "Extend"]

if "mapping_columns" not in st.session_state:
    st.session_state.mapping_columns = [
        {"export_name": name, "col_a": "", "col_b": ""} for name in DEFAULT_COLUMNS
    ]
if "df_a" not in st.session_state:
    st.session_state.df_a = None
    st.session_state.sheet_a = None
    st.session_state.file_a_name = None
if "df_b" not in st.session_state:
    st.session_state.df_b = None
    st.session_state.sheet_b = None
    st.session_state.file_b_name = None
if "output_dir" not in st.session_state:
    st.session_state.output_dir = os.getcwd()
if "merge_result" not in st.session_state:
    st.session_state.merge_result = None  # 存储合并后的DataFrame用于预览

CONFIG_PATH = "mapping_config.json"

# ---------- 辅助函数 ----------
def normalize_column_name(name):
    if not isinstance(name, str):
        return name
    return name.strip().lower()

def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
            if "mapping_columns" in config:
                loaded = config["mapping_columns"]
                # 确保锁定列存在
                fixed_present = [c["export_name"] for c in loaded if c["export_name"] in FIXED_COLUMNS]
                for name in FIXED_COLUMNS:
                    if name not in fixed_present:
                        loaded.append({"export_name": name, "col_a": "", "col_b": ""})
                st.session_state.mapping_columns = loaded
            if "output_dir" in config and os.path.exists(config["output_dir"]):
                st.session_state.output_dir = config["output_dir"]
        st.toast("✅ 配置加载成功", icon="📂")
    else:
        st.toast("未找到历史配置文件", icon="ℹ️")

def save_config():
    config = {
        "mapping_columns": st.session_state.mapping_columns,
        "output_dir": st.session_state.output_dir
    }
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    st.toast("💾 配置已保存", icon="✅")

def standardize(df, table):
    out = pd.DataFrame(index=df.index)
    norm_map = {normalize_column_name(col): col for col in df.columns}
    
    for cfg in st.session_state.mapping_columns:
        col = cfg["export_name"]
        src = cfg["col_a"] if table == "a" else cfg["col_b"]
        if src and src.strip():
            if src in df.columns:
                out[col] = df[src]
            else:
                norm_src = normalize_column_name(src)
                if norm_src in norm_map:
                    out[col] = df[norm_map[norm_src]]
                else:
                    out[col] = None
        else:
            out[col] = None
    
    # 数据来源
    if table == "a":
        fname = os.path.splitext(os.path.basename(st.session_state.file_a_name))[0]
        sheet = st.session_state.sheet_a
        out["数据来源"] = f"{fname}({sheet})"
    else:
        fname = os.path.splitext(os.path.basename(st.session_state.file_b_name))[0]
        sheet = st.session_state.sheet_b
        out["数据来源"] = f"{fname}({sheet})"
    return out

def perform_merge_and_validation():
    """执行合并与校验，返回结果DataFrame和冲突行数"""
    if st.session_state.df_a is None or st.session_state.df_b is None:
        st.error("请先上传表A和表B")
        return None, 0
    
    has_key = any(c["export_name"] == "客户料号" and (c["col_a"] or c["col_b"]) for c in st.session_state.mapping_columns)
    has_price = any(c["export_name"] == "单价" and (c["col_a"] or c["col_b"]) for c in st.session_state.mapping_columns)
    if not has_key:
        st.error("客户料号必须映射有效列")
        return None, 0
    if not has_price:
        st.error("单价必须映射有效列")
        return None, 0

    df_a_std = standardize(st.session_state.df_a, "a")
    df_b_std = standardize(st.session_state.df_b, "b")

    # 内部冲突检测
    def internal_conflicts(df):
        conflicts = set()
        price_map = {}
        for idx, row in df.iterrows():
            pn = row["客户料号"]
            price = row["单价"]
            if pd.isna(pn) or pd.isna(price):
                continue
            if pn not in price_map:
                price_map[pn] = price
            elif abs(price_map[pn] - price) > 0.0001:
                conflicts.add(pn)
        return conflicts

    a_internal = internal_conflicts(df_a_std)
    b_internal = internal_conflicts(df_b_std)

    a_price = df_a_std.dropna(subset=["单价"]).groupby("客户料号")["单价"].first().to_dict()
    b_price = df_b_std.dropna(subset=["单价"]).groupby("客户料号")["单价"].first().to_dict()
    cross_conflicts = set()
    for pn in a_price:
        if pn in b_price and abs(a_price[pn] - b_price[pn]) > 0.0001:
            cross_conflicts.add(pn)

    all_conflicts = a_internal | b_internal | cross_conflicts

    # 合并数据
    df_merged = pd.concat([df_a_std, df_b_std], ignore_index=True)
    df_merged["单价"] = pd.to_numeric(df_merged["单价"], errors="coerce")
    df_merged["是否单价冲突"] = df_merged["客户料号"].isin(all_conflicts)

    # 添加表A/表B原始单价列（用于展示对比）
    a_price_series = df_a_std.dropna(subset=["客户料号"]).drop_duplicates("客户料号").set_index("客户料号")["单价"]
    b_price_series = df_b_std.dropna(subset=["客户料号"]).drop_duplicates("客户料号").set_index("客户料号")["单价"]
    df_merged["表A原始单价"] = df_merged["客户料号"].map(a_price_series)
    df_merged["表B原始单价"] = df_merged["客户料号"].map(b_price_series)

    conflict_count = df_merged["是否单价冲突"].sum()
    return df_merged, conflict_count

# ---------- 页面标题（简洁）----------
st.markdown("""
<div style="margin-bottom: 0.5rem;">
    <h1 style="font-size: 2.2rem; font-weight: 700; color: #0f172a; margin-bottom: 0;">📊 高级表合并校验工具</h1>
    <p style="color: #64748b; font-size: 1rem;">合并表A与表B · 校验同料号不同单价 · 冲突高亮标黄</p>
</div>
""", unsafe_allow_html=True)

# ---------- 可折叠操作说明（轻量）----------
with st.expander("📖 操作步骤", expanded=False):
    st.markdown("""
    1. 上传表A和表B，选择对应工作表  
    2. 配置字段映射（**客户料号**与**单价**为锁定列，不可删除）  
    3. 设置输出目录  
    4. 点击 **合并并校验单价** 查看结果  
    5. 下载标黄冲突行的 Excel 文件
    """)

# ---------- 1. 上传区域（双栏卡片）----------
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📁 上传文件 <span>表A · 表B</span></div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="upload-col">', unsafe_allow_html=True)
    uploaded_a = st.file_uploader("拖拽或点击上传表A", type=["xlsx", "xls"], key="upload_a", label_visibility="collapsed")
    if uploaded_a:
        st.session_state.file_a_name = uploaded_a.name
        sheets_a = pd.ExcelFile(uploaded_a).sheet_names
        if len(sheets_a) == 1:
            st.session_state.sheet_a = sheets_a[0]
        else:
            st.session_state.sheet_a = st.selectbox("工作表", sheets_a, key="sheet_a_sel")
        st.session_state.df_a = pd.read_excel(uploaded_a, sheet_name=st.session_state.sheet_a)
        st.success(f"✅ 已加载 {uploaded_a.name} · {len(st.session_state.df_a)} 行")
    else:
        st.info("等待上传表A...")
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="upload-col">', unsafe_allow_html=True)
    uploaded_b = st.file_uploader("拖拽或点击上传表B", type=["xlsx", "xls"], key="upload_b", label_visibility="collapsed")
    if uploaded_b:
        st.session_state.file_b_name = uploaded_b.name
        sheets_b = pd.ExcelFile(uploaded_b).sheet_names
        if len(sheets_b) == 1:
            st.session_state.sheet_b = sheets_b[0]
        else:
            st.session_state.sheet_b = st.selectbox("工作表", sheets_b, key="sheet_b_sel")
        st.session_state.df_b = pd.read_excel(uploaded_b, sheet_name=st.session_state.sheet_b)
        st.success(f"✅ 已加载 {uploaded_b.name} · {len(st.session_state.df_b)} 行")
    else:
        st.info("等待上传表B...")
    st.markdown('</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ---------- 2. 字段映射配置（自定义表格）----------
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">🔧 字段映射配置</div>', unsafe_allow_html=True)

if st.session_state.df_a is not None and st.session_state.df_b is not None:
    cols_a = [""] + list(st.session_state.df_a.columns)
    cols_b = [""] + list(st.session_state.df_b.columns)

    # 确保锁定列存在
    current_mapping = st.session_state.mapping_columns.copy()
    fixed_present = [c["export_name"] for c in current_mapping if c["export_name"] in FIXED_COLUMNS]
    for name in FIXED_COLUMNS:
        if name not in fixed_present:
            current_mapping.append({"export_name": name, "col_a": "", "col_b": ""})

    # 表头
    st.markdown("""
    <div class="mapping-table-header">
        <div>导出列名</div>
        <div>表A实际列</div>
        <div>表B实际列</div>
        <div style="text-align:center">操作</div>
    </div>
    """, unsafe_allow_html=True)

    updated_mapping = []
    for idx, row in enumerate(current_mapping):
        export_name = row["export_name"]
        is_locked = export_name in FIXED_COLUMNS

        c1, c2, c3, c4 = st.columns([2.5, 2, 2, 0.8])
        with c1:
            if is_locked:
                st.text_input("导出列名", value=export_name, key=f"exp_{idx}", disabled=True, label_visibility="collapsed")
            else:
                new_name = st.text_input("导出列名", value=export_name, key=f"exp_{idx}", label_visibility="collapsed", placeholder="列名")
                export_name = new_name
        with c2:
            col_a_val = row["col_a"] if row["col_a"] in cols_a else ""
            new_col_a = st.selectbox("表A列", cols_a, index=cols_a.index(col_a_val) if col_a_val in cols_a else 0, key=f"cola_{idx}", label_visibility="collapsed")
        with c3:
            col_b_val = row["col_b"] if row["col_b"] in cols_b else ""
            new_col_b = st.selectbox("表B列", cols_b, index=cols_b.index(col_b_val) if col_b_val in cols_b else 0, key=f"colb_{idx}", label_visibility="collapsed")
        with c4:
            if is_locked:
                st.markdown('<div style="text-align:center"><span class="locked-badge">🔒 不可删</span></div>', unsafe_allow_html=True)
            else:
                if st.button("🗑️", key=f"del_{idx}", help="删除此行"):
                    continue  # 跳过此行即删除
        # 保存行（未被删除）
        if export_name.strip():
            updated_mapping.append({
                "export_name": export_name.strip(),
                "col_a": new_col_a if new_col_a != "" else "",
                "col_b": new_col_b if new_col_b != "" else ""
            })

    # 更新 session_state
    if updated_mapping:
        st.session_state.mapping_columns = updated_mapping

    # 添加自定义映射列按钮
    if st.button("➕ 添加自定义映射列", use_container_width=False):
        st.session_state.mapping_columns.append({"export_name": "", "col_a": "", "col_b": ""})
        st.rerun()

else:
    st.info("请先上传表A和表B，然后配置字段映射")
st.markdown('</div>', unsafe_allow_html=True)

# ---------- 3. 输出目录与操作按钮 ----------
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📂 输出设置 & 操作</div>', unsafe_allow_html=True)

c_dir1, c_dir2, c_dir3 = st.columns([5, 1, 1])
with c_dir1:
    dir_input = st.text_input("输出文件夹路径", st.session_state.output_dir, label_visibility="collapsed", key="dir_input")
with c_dir2:
    if st.button("📁 当前目录", use_container_width=True):
        st.session_state.output_dir = os.getcwd()
        st.rerun()
with c_dir3:
    if st.button("🔄 应用", use_container_width=True):
        if os.path.exists(dir_input):
            st.session_state.output_dir = dir_input
            st.toast("目录已更新", icon="✅")
        else:
            st.error("路径不存在")
os.makedirs(st.session_state.output_dir, exist_ok=True)

col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
with col_btn1:
    st.markdown('<div class="secondary-btn">', unsafe_allow_html=True)
    if st.button("💾 保存配置", use_container_width=True):
        save_config()
    st.markdown('</div>', unsafe_allow_html=True)
with col_btn2:
    st.markdown('<div class="secondary-btn">', unsafe_allow_html=True)
    if st.button("📂 加载配置", use_container_width=True):
        load_config()
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
with col_btn3:
    st.markdown('<div class="primary-gradient-btn">', unsafe_allow_html=True)
    merge_clicked = st.button("🚀 合并并校验单价", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# ---------- 4. 合并结果预览区 ----------
if merge_clicked:
    with st.spinner("正在合并并校验..."):
        result_df, conflict_cnt = perform_merge_and_validation()
        st.session_state.merge_result = result_df

if st.session_state.merge_result is not None:
    df = st.session_state.merge_result
    total_rows = len(df)
    conflict_rows = df["是否单价冲突"].sum()

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">📋 合并结果预览</div>', unsafe_allow_html=True)

    # 统计徽章
    st.markdown(f"""
    <div style="margin-bottom: 16px;">
        <span class="stats-badge">📊 总匹配行数: {total_rows}</span>
        <span class="stats-badge" style="background: #fef9c3; border-color: #eab308;">⚠️ 冲突行数: {conflict_rows}</span>
    </div>
    """, unsafe_allow_html=True)

    # 表格展示
    st.dataframe(df, use_container_width=True, height=400)
    st.caption("💡 下载的Excel文件中冲突行将自动标黄高亮")

    # 下载按钮
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_filename = f"合并校验结果_{now}.xlsx"
    out_path = os.path.join(st.session_state.output_dir, out_filename)

    # 写入Excel并标黄
    df.to_excel(out_path, index=False, engine="openpyxl")
    wb = load_workbook(out_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    headers = [cell.value for cell in ws[1]]
    conflict_col_idx = headers.index("是否单价冲突") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=conflict_col_idx).value is True:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
    wb.save(out_path)

    with open(out_path, "rb") as f:
        st.download_button(
            label="📥 下载标黄冲突Excel",
            data=f,
            file_name=out_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    st.success(f"文件已保存至: {out_path}")
    st.markdown('</div>', unsafe_allow_html=True)
